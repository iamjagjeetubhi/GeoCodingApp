from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, update_session_auth_hash
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth.forms import AdminPasswordChangeForm
from django.core.mail import EmailMessage
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render, get_list_or_404, get_object_or_404
from django.template.loader import render_to_string
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.translation import ugettext_lazy as _
from .forms import SignupForm, UserForm, PasswordChangeForm, FileUploadForm
from .tokens import account_activation_token
from .models import User
from django.utils.decorators import method_decorator
from django.conf import settings

def signup(request):
    if request.user.is_authenticated():
        return redirect('view_profile')
    else:
        if request.method == 'POST':
            form = SignupForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                user.is_active = False
                user.save()
                current_site = get_current_site(request)
                subject = 'Activate your accounts.'
                message = render_to_string('accounts/activation_email.html', {
                    'user':user, 'domain':current_site.domain,
                    'uid': urlsafe_base64_encode(force_bytes(user.pk)),
                    'token': account_activation_token.make_token(user),
                })
                toemail = form.cleaned_data.get('email')
                email = EmailMessage(subject, message, to=[toemail])
                email.send()
                return render(request, 'accounts/activation_pending.html')
        else:
            form = SignupForm()
        return render(request, 'registration/register.html', {'form': form})


def activate(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except(TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    if user is not None and account_activation_token.check_token(user, token):
        user.is_active = True
        user.backend='django.contrib.auth.backends.ModelBackend'
        user.save()
        login(request, user)
        return render(request, 'accounts/activation_completed.html')
    else:
        return HttpResponse('Activation link is invalid!')

def index(request):
    if request.user.is_authenticated:
        return redirect('userpage')
    else:
        return redirect('login')

@login_required
def userpage(request):
    username = None
    if request.user.country:
       return redirect('view_profile')
    else:
       return redirect('update_profile')

import requests
def get_google_results(address, api_key=None):
    """
    Get geocode results from Google Maps Geocoding API.
    
    Note, that in the case of multiple google geocode reuslts, this function returns details of the FIRST result.
    
    @param address: String address as accurate as possible. For Example "Maha Singh Nagar, Daba Lohara Road, Ludhiana"
    @param api_key: String API key if present from google. 
                    If supplied, requests will use your allowance from the Google API. If not, you
                    will be limited to the free usage of 2500 requests per day.
    """
    # Set up your Geocoding url
    geocode_url = "https://maps.googleapis.com/maps/api/geocode/json?address={}".format(address)
    if api_key is not None:
        geocode_url = geocode_url + "&key={}".format(api_key)
        
    # Ping google for the reuslts:
    results = requests.get(geocode_url)
    # Results will be in JSON format - convert to dict using requests functionality
    results = results.json()
    
    # if there's no results or an error, return empty results.
    if len(results['results']) == 0:
        output = {
            "latitude": "Can't Fetch",
            "longitude": "Can't Fetch",
        }
    else:    
        answer = results['results'][0]
        output = {
            "latitude": answer.get('geometry').get('location').get('lat'),
            "longitude": answer.get('geometry').get('location').get('lng'),
        }
        
    return output

@login_required
def view_profile(request):
    if request.method == 'POST':
        upload_form = FileUploadForm(request.POST, request.FILES, instance=request.user)
        if upload_form.is_valid():
            upload_form.save()
            print(settings.MEDIA_ROOT+"/"+request.user.upload.name)
            import openpyxl
            book = openpyxl.load_workbook(settings.MEDIA_ROOT+"/"+request.user.upload.name)
            sheet = book.active
            column_count = sheet.max_column
            for index,row in enumerate(sheet.rows):
                for cell in row:
                    #print(cell.value)
                    result = get_google_results(cell.value, "AIzaSyC88h272ksL-Wl09gAIz_Zuqsb4DqTjHmY")
                    #print(result)
                sheet.cell(row=index+1,\
                          column=column_count+1)\
                          .value= "lat / "+ str(result["latitude"])
                sheet.cell(row=index+1,\
                          column=column_count+2)\
                          .value= "lon / "+ str(result["longitude"])
            book.save(settings.MEDIA_ROOT+"/"+request.user.upload.name)
            messages.success(request, _('Your profile was successfully updated!'))
            return redirect('view_profile')
        else:
            messages.error(request, _('Please correct the error below.'))
    else:
        upload_form = FileUploadForm(instance=request.user)
    return render(request, 'accounts/view_profile.html', {
        'upload_form':upload_form,
        })

@login_required
@transaction.atomic
def update_profile(request):
    if request.method == 'POST' and 'profile' in request.POST:
        user_form = UserForm(request.POST, instance=request.user)
        if user_form.is_valid():
            user_form.save()
            messages.success(request, _('Your profile was successfully updated!'))
            return redirect('view_profile')
        else:
            messages.error(request, _('Please correct the error below.'))
    else:
        user_form = UserForm(instance=request.user)
    return render(request, 'accounts/edit_profile.html', {
        'user_form': user_form,
    })

@login_required
@transaction.atomic
def change_password(request):
    if request.user.has_usable_password():
        PasswordForm = PasswordChangeForm
    else:
        PasswordForm = AdminPasswordChangeForm
    if request.method == 'POST':
        form = PasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('view_profile')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordForm(request.user)
    return render(request, 'registration/password.html', {
        'form': form
    })

